"""
/*
This Python script is designed to elegantly extract Twitter usernames from a particular column in a spreadsheet containing the Twitter URLs.
Using  regex, it skillfully identifies Twitter handles embedded within URLs and provides a streamlined solution for gathering the usernames and pasting in a specified column.
Whether you're analyzing social media data or parsing user profiles, 
Simply supply the name of your excel file , the sheet name containing your Twitter URLs and the column you want to paste your usernames preferably the first column,
and let the script do the rest, delivering you the extracted Twitter usernames with precision.
**/
"""

import openpyxl
import re

# Load the Excel workbook
workbook = openpyxl.load_workbook('cs4.xlsx')

# Access the worksheet with the links
links_worksheet = workbook['T2']  # the naem of the sheet containing the twitter URLS

# Column indices for URLs and the column to paste extracted text
url_column_index = 3  # the column containing the URLS
extracted_text_column_index = 1  # the column to paste the extracted usernames



current_row = 2  # we start from 2 since there are headers

for row in links_worksheet.iter_rows(min_row=2, values_only=True):
    Twitter_links = row[3]  # Assuming the links are in the third column (column D)
    
    if Twitter_links is not None:
    # if Twitter_links:
        # Extract text from the URL
        start_index = Twitter_links.find("com/") + len("com/")
        match = re.search(r'[ ?]', Twitter_links[start_index:])
        
        if match:
            end_index = start_index + match.start()
            extracted_text = Twitter_links[start_index:end_index]
        else:
            extracted_text = Twitter_links[start_index:]
        
        # Print the extracted text for debugging
        print(extracted_text)

        # Write the extracted text to the first column of the current row
        links_worksheet.cell(row=current_row, column=1, value=extracted_text)

    current_row += 1  # Move to the next row
# Save the modified workbook
workbook.save('cs4.xlsx')
