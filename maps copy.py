import requests
import openpyxl

def address_to_zip(address):
    api_key = 'YOUR_API_KEY'  # Replace with your own API key
    base_url = 'https://maps.googleapis.com/maps/api/geocode/json'
    params = {
        'address': address,
        'key': api_key,
    }

    response = requests.get(base_url, params=params)

    if response.status_code == 200:
        data = response.json()
        if data.get('status') == 'OK':
            results = data.get('results', [])
            if results:
                zip_code = None
                for component in results[0].get('address_components', []):
                    if 'postal_code' in component.get('types', []):
                        zip_code = component.get('long_name')
                        break
                return zip_code

    return None

# Example usage:
# address = "555 7th St Santa Monica CA"
address = "MONROE TOWNSHIP"
zip_code = address_to_zip(address)

# ==================opens the excel work book========================
address_workbook = openpyxl.load_workbook('m.xlsx')

address_sheetName = address_workbook['Sheet1']

addresses_list = []

# Iterate over the rows in the worksheet and extract the links
# for row in address_sheetName.iter_rows(min_row=2, values_only=True):
#     addresses_column = row[0]  # column of the addresses
#     if addresses_column is not None:
#         addresses_list.append(addresses_column)
#         # print(addresses_column)
#         zipcode = address_to_zip(addresses_column)
        # print(f'{addresses_column} Zip code is {zipcode}')

# Iterate over the rows in the worksheet and extract the links
for row_idx, row in enumerate(address_sheetName.iter_rows(min_row=2, values_only=True), start=2):
    addresses_column = row[0]  # column of the addresses
    if addresses_column is not None:
        zipcode = address_to_zip(addresses_column)
        print(f'{addresses_column} Zip code is {zipcode}')
        if zipcode:
            address_sheetName.cell(row=row_idx, column=2, value=zipcode)

# Save the updated workbook
address_workbook.save('m.xlsx')




# for add in addresses_list:
#     print(add)


# ===================================================

# if zip_code:
#     print("ZIP Code:", zip_code)
# else:
#     print("ZIP Code not found for the given address.")
