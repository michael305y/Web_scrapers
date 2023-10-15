# import requests
# from bs4 import BeautifulSoup

# url = "https://www.instagram.com/repti7ian.luxury/"

# response = requests.get(url)
# if response.status_code == 200:
#     soup = BeautifulSoup(response.content, "html.parser")
    
#     # Extract profile information
#     profile_name = soup.find("h1", {"class": "rhpdm"}).text
#     followers_count = soup.find("span", {"class": "g47SY"}).text
#     posts_count = soup.find_all("span", {"class": "g47SY"})[0].text
#     following_count = soup.find_all("span", {"class": "g47SY"})[1].text

#     print("Profile Name:", profile_name)
#     print("Followers:", followers_count)
#     print("Posts:", posts_count)
#     print("Following:", following_count)
# else:
#     print("Failed to retrieve the page.")

# =================================================================================================

# # url = "https://www.instagram.com/repti7ian.luxury/"
# url = "https://www.instagram.com/dianadimeo_/?ig_rid=79be9af1-cb56-445b-a109-11774e6faac3"


# start_index = url.rfind("com/") + len("com/")
# end_index = url.find("/", start_index)

# extracted_text = url[start_index:end_index]
# print(extracted_text)

# ============================================================================

import openpyxl

# Load the Excel workbook
workbook = openpyxl.load_workbook('cs2.xlsx')

# Choose the specific sheet within the workbook
# sheet = workbook.active

# Access the worksheet with the links
links_worksheet = workbook['IG'] 

# Column indices for URLs and the column to paste extracted text
url_column_index = 3  # Adjust as needed
extracted_text_column_index = 1  # Adjust as needed

# Create a list to store the links
links = []

# Iterate over the rows in the worksheet and extract the links
for row in links_worksheet.iter_rows(min_row=2, values_only=True):
    IG_links = row[3]  # Assuming the links are in the third column based on index numbering (column D)
    if IG_links is not None:
        links.append(IG_links)

# for testing if the links agre being retrieved
for link in links:
    print(link)
    
        # Extract text from the URL
    start_index = link.rfind("com/") + len("com/")
    end_index = link.find("/", start_index)
    extracted_text = link[start_index:end_index]

    print(extracted_text)

    links_worksheet.cell(row=links.index(link) + 2, column=1, value=extracted_text)

            # Write extracted text to the corresponding cell in the new column
        # extracted_text_cell = links_worksheet.cell(row=row[0].row, column=extracted_text_column_index)
        # extracted_text_cell.value = extracted_text

# Save the modified workbook
workbook.save('cs3.xlsx')

# # Loop through rows in the sheet
# for row in links_worksheet.iter_rows(min_row=2, max_row=links_worksheet.max_row, min_col=url_column_index, max_col=url_column_index):
#     url = row[0].value
    
#     # Check if the cell is not empty
#     if url:
#         # Extract text from the URL
#         start_index = url.rfind("com/") + len("com/")
#         end_index = url.find("/", start_index)
#         extracted_text = url[start_index:end_index]

#         print(extracted_text)
        
        # Write extracted text to the corresponding cell in the new column
        # extracted_text_cell = sheet.cell(row=row[0].row, column=extracted_text_column_index)
        # extracted_text_cell.value = extracted_text

# Save the modified workbook
# workbook.save('Content creators.xlsx')

