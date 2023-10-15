'''
Script to automate the process of retrieving ZIP codes from a dataset containing various addresses.
'''

import time
import openpyxl
from geopy.geocoders import Nominatim 

# Initialize Nominatim API 
geolocator = Nominatim(user_agent="geoapiExercises") 

addresses_not_found = []  # to collect all addresses that have no zip code

# ==================opening and loading the excel work book========================
try:
     address_workbook = openpyxl.load_workbook('m.xlsx')
     address_sheetName = address_workbook['Sheet1']  #specifying the sheet conataining the addresses
except FileNotFoundError as error:
    print(f'{error} File not found, please check the name/path of the file')
    exit()

# Iterate over the rows in the worksheet and extract the addressed
for row_idx, column in enumerate(address_sheetName.iter_rows(min_row=2, values_only=True), start=2):
    addresses_column = column[0]  # column containing the addresses based on index(e.g. 0 will be column A)
    if addresses_column is not None:
        try:
            zipcode = geolocator.geocode(addresses_column)
            if zipcode:
                data = zipcode.raw 
                loc_data = data['display_name'].split()
                # print("Full Location") # for testing purposes
                # print(loc_data) 
                # print(zip)
                # print("Zip code : ",loc_data[-3])
                code = loc_data[-3].rstrip(',')  # index position of the zipcode in the location list
                print(f'{addresses_column} Zip code is {code}')
                address_sheetName.cell(row=row_idx, column=2, value=code) # if zipcode is found
            else:
                print(f'{addresses_column} Zip code not found')
                address_sheetName.cell(row=row_idx, column=2, value='NOT FOUND')  #if no zip code is found
                addresses_not_found.append(addresses_column)
        except Exception as e:
            print(f'Error geocoding {addresses_column}: {e}')


No_of_addresses_Not_found = len(addresses_not_found)
print('gathering all missing zipcodes... please wait')
time.sleep(2)
print('check done...')
time.sleep(1)
print(f'Only {No_of_addresses_Not_found}' + 'zipcodes not found')
print(addresses_not_found)
print('printing all the addresses not found....')
time.sleep(2)
for number, address in enumerate(addresses_not_found, start=1):
    print(f'{number}: {address}')


# # saving the workbook after processing
address_workbook.save('m.xlsx')
address_workbook.close()
