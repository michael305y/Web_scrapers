### using selenium to automate the web
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.keys import Keys

from selenium.webdriver.support.ui import Select

###
from webdriver_manager.chrome import ChromeDriverManager
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
###

# # Create a Service object
# service = webdriver.chrome.service.Service(ChromeDriverManager().install())

# # Pass in the Service object to the webdriver.Chrome method
# driver = webdriver.Chrome(service=service)

# driver.get("https://twitter.com/alisonrosejeff1")
# driver.get("https://www.instagram.com/pepamack/")

# # # Add a delay to allow the page to load
# time.sleep(2)

#  # Wait for the element to be visible
# wait = WebDriverWait(driver, 5)
# element = wait.until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[2]/div/div/div[2]/div/div/div/div[1]/div[2]/section/nav/div[2]/div/div/div[2]/div/div')))

# IG_name = driver.find_element(By.XPATH, '/html/body/div[2]/div/div/div[2]/div/div/div/div[1]/div[2]/section/main/div/header/section/div[1]/a/h2')

# # print(IG_search.send_keys("lllll"))

# print(IG_name.text)

# follower_count = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[2]/main/div/div/div/div[1]/div/div[3]/div/div/div/div/div[5]/div[2]/a/span[1]/span')

# print(follower_count.text)

# time.sleep(3)

###### Method 2 Getting data from Excel files ######
# # Load the Excel file       D:\Project Apps\Selenium_scripts\Female Content creators.xlsx
workbook = openpyxl.load_workbook('Female_Content_ss.xlsx')

# Access the worksheet with the links
links_worksheet = workbook['Twitter'] 

# Create a new sheet for the results
results_sheet = workbook.create_sheet("Results")
results_sheet.append(["User Name", "Follower Count"])  # Write header row

# Create a list to store the links
links = []

# Iterate over the rows in the worksheet and extract the links
for row in links_worksheet.iter_rows(min_row=2, values_only=True):
    IG_links = row[3]  # Assuming the links are in the first column (column D)
    if IG_links is not None:
        links.append(IG_links)

# for testing if the links agre being retrieved
for link in links:
    print(link)

# =======================================================
# Create a Service object
service = webdriver.chrome.service.Service(ChromeDriverManager().install())

# Pass in the Service object to the webdriver.Chrome method
driver = webdriver.Chrome(service=service)

# # Loop through the links and perform operations
for link in links:
    driver.get(link)

    time.sleep(2)

    #     # Wait for the element to be visible
    wait = WebDriverWait(driver, 5)
    element = wait.until(EC.visibility_of_element_located((By.XPATH, '/html/body/div[1]/div/div/div[2]/main/div/div/div/div[1]/div/div[3]/div/div/div/div/div[5]/div[2]/a/span[1]/span')))
    
    # tiktok_name_element = driver.find_element(By.XPATH, '/html/body/div[1]/div[2]/div[2]/div/div[1]/div[1]/div[2]/h1')
    twitter_follower_element = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/div[2]/main/div/div/div/div[1]/div/div[3]/div/div/div/div/div[5]/div[2]/a/span[1]/span')
    
    # Extract text from the WebElement
    # tiktok_name_text = tiktok_name_element.text
    twitter_follower_text = twitter_follower_element.text

    # results_sheet.cell(row=links.index(link) + 2, column=1, value=tiktok_name_text)
    results_sheet.cell(row=links.index(link) + 2, column=2, value=twitter_follower_text)
    # results_sheet.append([IG_name_text])  # Append to the results sheet

    time.sleep(3)

    print(twitter_follower_text)

# Save the changes to the workbook
workbook.save('Female Content creators.xlsx')

# Close the workbook
workbook.close()




############///////

#     # wait 2s before addig the next row
#     time.sleep(2)


    ###########
    # optional to check if the message success displayed within 10s if not it retries
    # while True:
    #     # code to submit the form
    #     # button to save and insert items into the DB
    #     save_button = driver.find_element(By.XPATH, "/html/body/div[1]/div[2]/div/div[1]/div/form/div/div/input[2]")
    #     save_button.click()

    #     # wait for the success message to appear
    #     try:
    #         success_message = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/div[2]/div/ul/li')))
    #         print(success_message.text)
    #         break

    #     except:
            
    #         print("Submission failed. Retrying...")


    
while True:
    command = input("Enter Crtl + C to quit")
    # time.sleep(1)